#!/usr/bin/env python3
"""Parse CMU transcript and fill the Samsung 이수교과목 Excel template."""

from typing import Optional

from openpyxl import load_workbook

# CMU: 3 units = 1 semester credit
def units_to_credits(units: float) -> float:
    if units == 0:
        return 0
    return round(units / 3, 2)

# Course data extracted from CMU transcript (Dongyoon Kim)
# Format: (course_num, title, dept, units, grade, year, semester, is_retake)
# Semester: 1=Spring, 2=Fall, 3=Summer
COURSES = [
    # AP/Transfer
    ("21120", "DIFFERENTIAL INT CAL", "MSC", 10.0, "AD", None, "AP", False),
    ("21122", "INTEGRTN & APPROX", "MSC", 10.0, "AD", None, "AP", False),
    ("36201", "STATS REASON PRACTCE", "STA", 9.0, "AD", None, "AP", False),
    ("15110", "PRINCPLS OF COMPUTNG", "CS", 10.0, "A", None, "AP", False),
    ("21240", "MATRIX ALGBA APPLC", "MSC", 10.0, "A", None, "AP", False),
    ("33124", "INTRO TO ASTRONOMY", "PHY", 9.0, "B", None, "AP", False),
    ("36146", "FRESH SEM STATISTICS", "STA", 9.0, "A", None, "AP", False),
    ("67100", "INFO SYS FRSH WRKSHP", "ISH", 1.0, "P", None, "AP", False),
    ("79104", "GLOBAL HISTORIES", "HIS", 9.0, "B", None, "AP", False),
    ("99101", "CMPTNG CARNEGIE MELL", "CMU", 3.0, "P", None, "AP", False),
    # Fall 2016
    ("15112", "FNDMTLS OF PGMG & CS", "CS", 12.0, "B", 2016, "2", False),
    ("21256", "MULTIVARIATE ANALYS", "MSC", 9.0, "A", 2016, "2", False),
    ("36202", "STATISTICAL METHODS", "STA", 9.0, "A", 2016, "2", False),
    ("67250", "INFO SYSTEMS MILIEUX", "ISH", 9.0, "B", 2016, "2", False),
    ("76101", "INTERPRETN & ARGMNT", "ENG", 9.0, "B", 2016, "2", False),
    # Spring 2017
    ("15122", "PRIN IMPRTV COMPTATN", "CS", 10.0, "D", 2017, "1", False),
    ("36225", "INTRO PROBLTY THEORY", "STA", 9.0, "C", 2017, "1", False),
    ("67262", "DTABASE DSGN/DEV", "ISH", 9.0, "B", 2017, "1", False),
    ("82133", "ELE CHINESE ONLINE I", "ML", 12.0, "A", 2017, "1", False),
    # Fall 2017
    ("21127", "CONCEPTS OF MATHMTCS", "MSC", 10.0, "C", 2017, "2", False),
    ("36226", "INTRO STATCL INFRCE", "STA", 9.0, "B", 2017, "2", False),
    ("67272", "APPLCTN DESGN & DEV", "ISH", 9.0, "W", 2017, "2", False),  # Withdrawn
    ("67364", "PRACTICAL DATA SCI", "ISH", 9.0, "A", 2017, "2", False),
    ("80249", "AUTOMATIO AI SOCIETY", "PHI", 9.0, "A", 2017, "2", False),
    # Spring 2018
    ("36350", "STATISTICAL COMPUTNG", "STA", 9.0, "A", 2018, "1", False),
    ("36401", "MODERN REGRESSION", "STA", 9.0, "B", 2018, "1", False),
    ("57209", "THE BEATLES", "MUS", 9.0, "A", 2018, "1", False),
    ("67272", "APPLCTN DESGN & DEV", "ISH", 9.0, "C", 2018, "1", True),   # Retake
    ("88275", "BUBBLES: DATA SCI", "SDS", 9.0, "A", 2018, "1", False),
    # Fall 2018
    ("36315", "STAT GRPHCS & VISUAL", "STA", 9.0, "B", 2018, "2", False),
    ("36402", "ADV METHODS DATA ANL", "STA", 9.0, "C", 2018, "2", False),
    ("36410", "INTRO PROB MODELING", "STA", 9.0, "D", 2018, "2", False),
    ("36462", "SPC TP: DATA MINING", "STA", 9.0, "C", 2018, "2", False),
    ("67373", "IS CONSULTING PROJ", "ISH", 12.0, "A", 2018, "2", False),
    # Spring 2019
    ("67505", "INFO SYSTMS INTRNSHP", "ISH", 3.0, "P", 2019, "1", False),
    # Summer 2019
    ("05438", "RLE TCH LRNG 21ST CN", "HCI", 12.0, "B", 2019, "3", False),
    ("15294", "RAPID PROTYP TEC", "CS", 5.0, "B", 2019, "3", False),
    ("67309", "SPECIAL TOPICS", "ISH", 6.0, "B", 2019, "3", False),
    ("67442", "IOS ENGINEERING", "ISH", 9.0, "B", 2019, "3", False),
    ("70415", "INTRO ENTREPRENRSHIP", "BUS", 9.0, "B", 2019, "3", False),
    ("76270", "WRITNG FOR PROFSNS", "ENG", 9.0, "A", 2019, "3", False),
    # Fall 2019
    ("17200", "ETHCS POLICY CMPTNG", "ISR", 9.0, "B", 2019, "2", False),
    # Spring 2020 (last undergrad semester, start of grad)
    ("94700", "ORGNZTNL DSGN & IMPL", "HC", 6.0, "A", 2020, "1", False),
    ("95703", "DATABASE MANAGEMENT", "ISM", 12.0, "B", 2020, "1", False),
    ("95710", "ECONOMIC ANALYSIS", "ISM", 6.0, "A-", 2020, "1", False),
    ("95712", "OBJ ORIENT PRG JAVA", "ISM", 12.0, "C", 2020, "1", False),
    ("95715", "FINANCIAL ACCOUNTING", "ISM", 6.0, "B", 2020, "1", False),
    ("95717", "WRTG INF SYS MGT", "ISM", 6.0, "B+", 2020, "1", False),
    ("95796", "STATS FOR IT MANAGER", "ISM", 6.0, "B+", 2020, "1", False),
    # Fall 2020
    ("95702", "DIST SYSTEM ISM", "ISM", 12.0, "B-", 2020, "2", False),
    ("95716", "PRINCIPLS OF FINANCE", "ISM", 6.0, "B-", 2020, "2", False),
    ("95718", "PROFESSIONAL SPEAKNG", "ISM", 6.0, "A-", 2020, "2", False),
    ("95722", "DIGITAL TRANSFORMATN", "ISM", 6.0, "A-", 2020, "2", False),
    ("95723", "MANG DISRUPT TECH", "ISM", 6.0, "B+", 2020, "2", False),
    ("95760", "DCSN MKG UNCERTAINTY", "ISM", 6.0, "B-", 2020, "2", False),
    ("95851", "MAKING PRODCTS COUNT", "ISM", 6.0, "A", 2020, "2", False),
    ("95874", "AGILE METHODS", "ISM", 6.0, "A+", 2020, "2", False),
    # Spring 2021
    ("95900", "INTERNSHIP", "ISM", 0.0, "P", 2021, "1", False),
    # Summer 2021 - no courses
    # Fall 2021
    ("94800", "NEGOTIATION", "HC", 6.0, "B+", 2021, "2", False),
    ("95720", "INFO SYSTEMS PROJECT", "ISM", 18.0, "B+", 2021, "2", False),
    ("95729", "E-COM TCH ML AN BOTS", "ISM", 6.0, "B", 2021, "2", False),
    ("95733", "INTERNET OF THINGS", "ISM", 6.0, "P", 2021, "2", False),
    ("95734", "MANGING DIGITAL BUS", "ISM", 12.0, "A", 2021, "2", False),
    ("95775", "IT BUSINESS LDRSHP", "ISM", 6.0, "A", 2021, "2", False),
]


def get_major(course: tuple) -> str:
    """Map course to 전공명. ISM/HC (grad) -> 경영정보시스템, else -> 수학통계학."""
    _, _, dept, _, _, _, _, _ = course
    if dept in ("ISM", "HC"):
        return "경영정보시스템"
    return "수학통계학"


def get_semester_display(sem: str, year: Optional[int]) -> str:
    """Convert to Korean semester format."""
    if sem == "AP":
        return "인정"
    if sem == "1":
        return "1학기"
    if sem == "2":
        return "2학기"
    if sem == "3":
        return "여름학기"
    return ""


def main():
    excel_path = "/Users/johnkim/Downloads/이수교과목_20260310163546.xlsx"
    wb = load_workbook(excel_path)
    ws = wb["교과목이수 입력칸"]

    courses_to_add = COURSES

    # Write rows starting at row 2
    for i, course in enumerate(courses_to_add):
        row_idx = i + 2
        course_num, title, dept, units, grade, year, sem, is_retake = course
        full_title = f"{course_num} {title}"
        credits = units_to_credits(units)
        year_str = str(year) if year else ""
        sem_str = get_semester_display(sem, year)
        major = get_major(course)
        retake_str = "재수강" if is_retake else ""

        ws.cell(row=row_idx, column=1, value=i + 1)  # NO
        ws.cell(row=row_idx, column=2, value=major)  # 전공명
        ws.cell(row=row_idx, column=3, value=year_str)  # 수강연도
        ws.cell(row=row_idx, column=4, value=sem_str)  # 학기
        ws.cell(row=row_idx, column=5, value=full_title)  # 과목명
        ws.cell(row=row_idx, column=6, value=None)  # 과목유형 - leave blank
        ws.cell(row=row_idx, column=7, value=credits if credits > 0 else None)  # 취득학점
        ws.cell(row=row_idx, column=8, value=grade)  # 성적
        ws.cell(row=row_idx, column=9, value=retake_str)  # 재수강여부

    out_path = "/Users/johnkim/Downloads/이수교과목_20260310163546.xlsx"
    wb.save(out_path)
    print(f"Saved to {out_path}")
    print(f"Added {len(courses_to_add)} courses")


if __name__ == "__main__":
    main()
